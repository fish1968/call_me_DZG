
import openpyxl
from functions import judge, cankill_montecarlo, debug

filename = r"C:\Users\97738\Desktop\call_me_DZG_judger_for_HuBiaoTasks\劫镖计算\test.xlsx"
guess_names = []
guess_values = []

def sum_of_list_by_index(lis, a,b,c,d,e):
    return lis[a] + lis[b] + lis[c] + lis[d] + lis[e]
    



def obtain_mylist(excel_name, sheetname, names, values, debug = False):
    import os
    print(os.getcwd())
    workbook = openpyxl.load_workbook(excel_name)
    worksheet = workbook[sheetname]
    has_end = False
    for index in range(0, worksheet.max_row):
#		for (j, col) in enumerate(worksheet.iter_cols(1, worksheet.max_column)):
        if has_end == True: break
        for (j, col) in enumerate(worksheet.iter_cols(1, 2)):
            # if (len(col[i].value) <= 1): break
            try: 
                col[index]
                if (index> 0) and (j == 0):
                    
                    if (len(col[index].value) <= 1): has_end = True
                    names.append(col[index].value.strip(""))
                if (index> 0) and (j == 1):
                    values.append(col[index].value)
                if debug == True:
                    print(col[index].value, end="\t\t")
            except:
                has_end = True
            if has_end == True: break
                
        if debug == True: print('')

    if debug == False: return
    for (index,name) in enumerate(names):
        #print(names)
        #print(values)
        print(f" The {index+1}th person is {name} with a value of {values[index]}")
        

def translate_from_idx_to_val(option, values):
    ans = []
    for idx in option:
        ans.append(values[idx])
    return ans

def obtain_minimum_choice_old_version( total_power_list, defender, data_sum, values):
    total_power = list(set(total_power_list))
    total_power.sort()
    length = len(total_power)
    for i in range(length):
        cur_sum = total_power[i]
        options = data_sum.get(cur_sum, None)
        if options == None: 
            continue
        else:
            for option in options:
                if debug: print(f"The current option is {option}")
                if cankill_montecarlo(translate_from_idx_to_val(option, values), defender):
                    return option
    return -1 # No choice available
    
def get_second_data(values: list[int], sum_list: list[int],  power_count: dict(), data_dict: dict(),  debug = False):
    length = len(values)
    num_of_enemy = 5
    for a in range(length-num_of_enemy):
        for b in range(a+1, length):
            for c in range(b+1, length):
                for d in range(c+1, length):
                    for e in range(d+1, length):
                        if debug == True: print(f"nums {a} {b} {c} {d} {e}")
                        # obtain a sequence of order one by one
                        total_power = sum_of_list_by_index(values, a, b, c, d, e)
                        if data_dict.get(total_power, -1) == -1:
                            data_dict[total_power] = ((a,b,c,d,e),)
                            power_count[total_power] = 1
                            sum_list.append(total_power)
                        else:
                            data_dict[total_power] = tuple(list(data_dict.get(total_power)) ) + ((a,b,c,d,e),) # data_sum[100] = origin_data_at_100 + (new tuple) -> (old tuples, new tuple)
                            power_count[total_power] += 1
    sum_list.sort()
    if debug == True:
        print("The power list is ")
        print(f"sum_list : {sum_list}")
        print("\n"*3)
        print("Our data_dict is ")
        print(f"Data dict is {data_dict}")
        print("Our sum_data is ")
        print(f"sum_data is {power_count}")

def test_choice(my_values, enemy_values, test_num = 200, probability=0.96) -> bool:
    return cankill_montecarlo(my_values, enemy_values, testNum=test_num, probability=probability)

def test_maximum_choice(my_values: list[int], enemy_values: list[int], test_num: int = 200):
    temp_values = my_values.copy()
    temp_values.sort()
    return test_choice(temp_values[-5:], enemy_values=enemy_values, test_num=test_num)

def obtain_minimum_choice(my_values: list[int], my_names: list[int], power_count: list[int], my_data: dict(), enemy_values: list[int], probability = 0.96, debug = False):
    power_count.sort()
    length = len(power_count)
    if test_maximum_choice(my_values, enemy_values, test_num=200) == False:
        print("No chance of winning  :< :<")
        return False
    for i in range(length):
        cur_sum = power_count[i]
        options = my_data.get(cur_sum, None)
        if options == None:
            continue
        for option in options:
            if test_choice(translate_from_idx_to_val(option, my_values), enemy_values, test_num=200, probability=probability) == True:
                if debug == True:
                    print(f"This is your best choice {translate_from_idx_to_val(option, my_names)}")
                return option
    if debug == True: print(f"Error in code!!!!!")
    return False

def guess(excel_name, guess_value):
    guess_names = []
    guess_values = []
    workbook = openpyxl.load_workbook(filename = excel_name)
    #	print(workbook.sheetnames)
    #	worksheet = workbook['others'] # obtain all names of sheet in the workbook
    worksheet = workbook['enemy-sanrenhe']
#	print(worksheet['B2'].value) # test whether the value at B2 can be read
    for i in range(0, worksheet.max_row):
        for (j, col) in enumerate(worksheet.iter_cols(1, 2)):
            try:
                print(col[i].value, end="\t\t")
                if (col[i].value == None): break
                if (i> 0) and (j == 0):
                    guess_names.append(col[i].value.strip(""))
                if (i> 0) and (j == 1):
                    guess_values.append(col[i].value)
            except:
                break
        print('')
    length = len(guess_names)
    sum = 0
    total_power = 0
    sum_dict = dict()
    data_sum = dict()
    sum_list = []
    sum_power_list  = []
    for a in range(length):
        if (a+5) > length: break # end of the for loop
        for b in range(a+1, length):
            for c in range(b+1, length):
                for d in range(c+1, length):
                    for e in range(d+1, length):
                        if debug: print(f"nums {a} {b} {c} {d} {e}")
                        # obtain a sequence of order one by one
                        total_power = sum_of_list_by_index(guess_values, a, b, c, d, e)
                        if sum_dict.get(total_power, -1) == -1:
                            sum_dict[total_power] = 1
                            sum_list.append(total_power)
                            data_sum[total_power] = ((a,b,c,d,e),)
                        else:
                            sum_dict[total_power] += 1
                            data_sum[total_power] = tuple(list(data_sum.get(total_power)) ) + ((a,b,c,d,e),) # data_sum[100] = origin_data_at_100 + (new tuple) -> (old tuples, new tuple)
                        sum_power_list.append(total_power)
                        sum += 1
    if guess_value in sum_list:
        print(f"The guess result is {data_sum[guess_value]}")
        print("Which is in names as ")
        for option in data_sum[guess_value]:
            print(translate_from_idx_to_val(option, guess_names),  end = "\n\t")
    else:
        print("No such casers")		

def enemy_combination(enemy_sum: int, enemy_values: list[int] ,error_rate = 0.01) -> dict:
    enemy_num = len(enemy_values)
    ans_dict = dict()
    for a1 in range(0, enemy_num-4):
        for a2 in range(a1+1, enemy_num-3):
            for a3 in range(a2+1, enemy_num-2):
                for a4 in range(a3+1, enemy_num-1):
                    for a5 in range(a4+1, enemy_num):
                        temp_sum = sum_of_list_by_index(enemy_values, a1, a2, a3, a4, a5)
                        error = temp_sum - enemy_sum
                        if abs(error) <= abs(error_rate*enemy_sum):
                            if ans_dict.get(error, -1) == -1:
                                ans_dict[error] = (1, (a1,a2,a3,a4,a5),)
                            else:
                                ans_dict[error] = (ans_dict[error][0]+1,) + ans_dict[error][1:] +  ((a1,a2,a3,a4,a5),)
    if ans_dict.get(0, -1) != -1:
        ans_dict["has_found"] = True
    else:
        ans_dict["has_found"] = False
    return ans_dict

def main():
    # read the data from xlsx
    # iterate through every possible combination, and find which is the best and cost least amount of combat power
    # sort summed power from low to high, then compare
    sum_power_list = []
    data_sum = dict()
    sum_dict = dict()
    sum_list = []
#	values = [100,97, 34, 52, 23, 43]
#	values = [1016,613 ,546 ,364 ,314 ,268 ,199 ,196 ,173 ,172 ,167 ,164 ,161 ,156 ,144 ,123 ,100]
    names = []
    values = []
    obtain_mylist("test.xlsx", "Sheet2", names, values)
    sum = 0
    total_power = 0
    length = len(values)
    for a in range(length):
        if (a+5) > length: break # end of the for loop
        for b in range(a+1, length):
            for c in range(b+1, length):
                for d in range(c+1, length):
                    for e in range(d+1, length):
                        if debug: print(f"nums {a} {b} {c} {d} {e}")
                        # obtain a sequence of order one by one
                        total_power = sum_of_list_by_index(values, a, b, c, d, e)
                        if sum_dict.get(total_power, -1) == -1:
                            sum_dict[total_power] = 1
                            sum_list.append(total_power)
                            data_sum[total_power] = ((a,b,c,d,e),)
                        else:
                            sum_dict[total_power] += 1
                            data_sum[total_power] = tuple(list(data_sum.get(total_power)) ) + ((a,b,c,d,e),) # data_sum[100] = origin_data_at_100 + (new tuple) -> (old tuples, new tuple)
                        sum_power_list.append(total_power)
                        sum += 1
    # sort the sum value 
    sum_power_list = list(set(sum_power_list))
    sum_power_list.sort()
    print(len(sum_power_list)) # number of possible sum together
    
#	enemy_option =  [188, 195, 243, 217, 195]
#	enemy_option =  [1200, 345, 253, 2000, 525]
#	enemy_option = [139,152,166,199,368] 
    #  [299,315,505,1326,506] # failed best at 0.9 (0,1,2,3,4)
    
    enemy_option = [ 456,636,273,216,258]


    find = obtain_minimum_choice(sum_power_list, enemy_option , data_sum, values = values)
    print(f"Our enemy is {enemy_option}")
    if find == -1:
        print("No solution could be found :<")
        print(f"Can you kill {False}")
    else:
        print(f"My solution in index is {find}, that in name is  {translate_from_idx_to_val(option = find, values=values)}")
        print(f"Can you kill? {True}")
        print(f"You shoud choose {translate_from_idx_to_val(option=find, values=names)}")
        
    if True: 
        print(f"The total num of value is {sum}")
        print(f"The sum of list is {sum_power_list}")
        for i, j in data_sum.items():
            print(f"The total force is {i}, and they could be formed by {j} and the total num of j is {len(j)}")
        print(sum)



#main()
if __name__ == "__main__":
    guess(filename, 930)
    print(cankill_montecarlo([613, 546, 123, 364, 144], [ 500, 326, 284, 323, 295]))
    print(cankill_montecarlo([1060,622,550,530,528], [1221, 546, 529, 315, 91], probability=0.3, debug=True))
    print(cankill_montecarlo([1243, 431, 311, 213, 199], [261, 244, 310, 191, 202], probability=0.3, debug=True))
    
    


#names = []
#values = []
#obtain_mylist("test.xlsx", names, values)

#print(names)
#print(values)

