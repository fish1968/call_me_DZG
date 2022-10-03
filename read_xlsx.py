# template for reading xlsx file's first and second row to the end
# reference : https://linuxhint.com/read-excel-file-python/#:~:text=The%20read_excel()%20function%20of,in%20the%20variable%20named%20data.
import openpyxl

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("test.xlsx")

# Define variable to read the active sheet:
worksheet = wookbook.active

# Iterate the loop to read the cell values
names = list()
values = list()


for i in range(0, worksheet.max_row):
    for (j, col) in enumerate(worksheet.iter_cols(1, worksheet.max_column)):
        print(col[i].value, end="\t\t")
        if (i> 0) and (j == 0):
            names.append(col[i].value.strip(""))
        if (i> 0) and (j == 1):
            values.append(col[i].value)
    print('')
    
print(names)
print(values)

for (i,name) in enumerate(names):
	print(f" The {i+1}th person is {name} with a value of {values[i]}")

# test case results

#Name            OurData
#小师妹          1016
#苗疆女          613
#李师师          546
#役使            364
#棋士            314
#西域女子                266
#将门女          199
#小乞丐          182
#绣娘            173
#丫鬟            172
#草原女孩                167
#采药女          164
#医师            161
#书香女          156
#针线女          144
#戏子            123
#卖伞女          100
#['小师妹', '苗疆女', '李师师', '役使', '棋士', '西域女子', '将门女', '小乞丐', '绣娘', '丫鬟', '草原女孩', '采药女', '医师', '
#书香女', '针线女', '戏子', '卖伞女']
#[1016, 613, 546, 364, 314, 266, 199, 182, 173, 172, 167, 164, 161, 156, 144, 123, 100]
# The 1th person is 小师妹 with a value of 1016
# The 2th person is 苗疆女 with a value of 613
# The 3th person is 李师师 with a value of 546
# The 4th person is 役使 with a value of 364
# The 5th person is 棋士 with a value of 314
# The 6th person is 西域女子 with a value of 266
# The 7th person is 将门女 with a value of 199
# The 8th person is 小乞丐 with a value of 182
# The 9th person is 绣娘 with a value of 173
# The 10th person is 丫鬟 with a value of 172
# The 11th person is 草原女孩 with a value of 167
# The 12th person is 采药女 with a value of 164
# The 13th person is 医师 with a value of 161
# The 14th person is 书香女 with a value of 156
# The 15th person is 针线女 with a value of 144
# The 16th person is 戏子 with a value of 123
# The 17th person is 卖伞女 with a value of 100
